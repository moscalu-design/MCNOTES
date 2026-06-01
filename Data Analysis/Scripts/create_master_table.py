from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SCRIPT_DIR = Path(__file__).resolve().parent
HERE = SCRIPT_DIR if (SCRIPT_DIR / "MC_Note_Datebase.xlsx").exists() else SCRIPT_DIR.parent
MC_DB = HERE / "MC_Note_Datebase.xlsx"
BO_DB = HERE / "BO_Data_Projects.xlsx"
MASTER = HERE / "Master Table.xlsx"
QUERY_FILE = HERE / "Master Table Power Query.m"

MC_COLUMNS_B_TO_AC = [
    "Source",
    "Template",
    "Extraction",
    "MC_Note_Type",
    "File Name",
    "Operation Number",
    "Validation Date",
    "Author",
    "Document Page Count",
    "Page count before opinion",
    "Annex Page Count",
    "Text Before Opinions",
    "OPS",
    "GLO",
    "PJ",
    "RM",
    "OCCO",
    "JU",
    "ECON",
    "CFC",
    "EIF",
    "FI",
    "IG",
    "PMM",
    "SG",
    "GIS",
    "HR",
    "OTHER",
]

BO_COLUMN_AB = "Operation Team OPS/GLO Main Division Short Name"
BO_OUTPUT_COLUMN = "BO Author (OPS/GLO)"
BO_PIN_GNG_VALIDATION_DATE = "PIN/GNG Validation Date"
BO_AFS_VALIDATION_DATE = "Operation AFS Validation Date"
BO_VALIDATION_OUTPUT_COLUMN = "BO Validation Date"
BO_PJ_COLUMN = "TEAM PJ"
BO_RM_COLUMN = "TEAM RM"
BO_JU_COLUMN = "TEAM JU"
BO_ECON_COLUMN = "Operation Team SG Main Division Short Name"
BO_PJ_OUTPUT_COLUMN = "BO PJ"
BO_RM_OUTPUT_COLUMN = "BO RM"
BO_JU_OUTPUT_COLUMN = "BO JU"
BO_ECON_OUTPUT_COLUMN = "BO ECON"

BO_SERVICE_COLUMN_MAP = {
    BO_PJ_COLUMN: BO_PJ_OUTPUT_COLUMN,
    BO_RM_COLUMN: BO_RM_OUTPUT_COLUMN,
    BO_JU_COLUMN: BO_JU_OUTPUT_COLUMN,
    BO_ECON_COLUMN: BO_ECON_OUTPUT_COLUMN,
}

BO_SELECTED_COLUMNS = [
    "Operation",
    BO_PIN_GNG_VALIDATION_DATE,
    BO_AFS_VALIDATION_DATE,
    BO_COLUMN_AB,
    *BO_SERVICE_COLUMN_MAP.keys(),
]
BO_EXPANDED_OUTPUT_COLUMNS = [
    f"BO {BO_PIN_GNG_VALIDATION_DATE}",
    f"BO {BO_AFS_VALIDATION_DATE}",
    BO_OUTPUT_COLUMN,
    *(f"BO {column}" for column in BO_SERVICE_COLUMN_MAP),
]
BO_RAW_EXPANDED_COLUMNS = [
    f"BO {BO_PIN_GNG_VALIDATION_DATE}",
    f"BO {BO_AFS_VALIDATION_DATE}",
    BO_OUTPUT_COLUMN,
    *(f"BO {column}" for column in BO_SERVICE_COLUMN_MAP),
]

MASTER_COLUMN_ORDER = [
    "Source",
    "Template",
    "Extraction",
    "MC_Note_Type",
    "File Name",
    "Operation Number",
    "Validation Date",
    "Author",
    BO_VALIDATION_OUTPUT_COLUMN,
    BO_OUTPUT_COLUMN,
    "Document Page Count",
    "Page count before opinion",
    "Annex Page Count",
    "Text Before Opinions",
    "OPS",
    "GLO",
    "PJ",
    "RM",
    "OCCO",
    "JU",
    "ECON",
    "CFC",
    "EIF",
    "FI",
    "IG",
    "PMM",
    "SG",
    "GIS",
    "HR",
    "OTHER",
    *BO_SERVICE_COLUMN_MAP.values(),
]


def m_string(value: Path | str) -> str:
    return '"' + str(value).replace('"', '""') + '"'


def m_list(values: list[str]) -> str:
    return "{" + ", ".join('"' + value.replace('"', '""') + '"' for value in values) + "}"


def build_power_query_formula() -> str:
    return f"""let
    PromoteHeaderByMarkers = (InputTable as table, Markers as list) as table =>
        let
            Rows = Table.ToRows(InputTable),
            TextRows = List.Transform(
                Rows,
                each List.Transform(_, each try Text.Trim(Text.From(_)) otherwise "")
            ),
            Scores = List.Transform(
                TextRows,
                each List.Count(List.Intersect({{_, Markers}}))
            ),
            BestScore = List.Max(Scores),
            HeaderIndex = if BestScore = 0 then 0 else List.PositionOf(Scores, BestScore),
            FromHeader = Table.Skip(InputTable, HeaderIndex),
            Promoted = Table.PromoteHeaders(FromHeader, [PromoteAllScalars=true])
        in
            Promoted,

    MCSource = Excel.Workbook(File.Contents({m_string(MC_DB)}), null, true),
    MCDatabaseRaw = MCSource{{[Item="Database", Kind="Sheet"]}}[Data],
    MCPromotedHeaders = PromoteHeaderByMarkers(
        MCDatabaseRaw,
        {{"Source", "Template", "Extraction", "File Name", "Operation Number"}}
    ),
    MCSelectedColumns = Table.SelectColumns(MCPromotedHeaders, {m_list(MC_COLUMNS_B_TO_AC)}, MissingField.UseNull),
    MCOperationTyped = Table.TransformColumns(
        MCSelectedColumns,
        {{{{"Operation Number", each try Int64.From(_) otherwise null, Int64.Type}}}}
    ),

    BOSource = Excel.Workbook(File.Contents({m_string(BO_DB)}), null, true),
    BODataRaw = BOSource{{[Item="BO Data", Kind="Sheet"]}}[Data],
    BOPromotedHeaders = PromoteHeaderByMarkers(
        BODataRaw,
        {{"Operation", "Operation Name", "{BO_COLUMN_AB}"}}
    ),
    BOSelectedColumns = Table.SelectColumns(
        BOPromotedHeaders,
        {m_list(BO_SELECTED_COLUMNS)},
        MissingField.UseNull
    ),
    BOOperationTyped = Table.TransformColumns(
        BOSelectedColumns,
        {{{{"Operation", each try Int64.From(_) otherwise null, Int64.Type}}}}
    ),
    BODeduped = Table.Distinct(BOOperationTyped, {{"Operation"}}),
    JoinedBO = Table.NestedJoin(
        MCOperationTyped,
        {{"Operation Number"}},
        BODeduped,
        {{"Operation"}},
        "BO",
        JoinKind.LeftOuter
    ),
    ExpandedBO = Table.ExpandTableColumn(
        JoinedBO,
        "BO",
        {m_list(BO_SELECTED_COLUMNS[1:])},
        {m_list(BO_EXPANDED_OUTPUT_COLUMNS)}
    ),
    BOValidationDate = Table.AddColumn(
        ExpandedBO,
        "{BO_VALIDATION_OUTPUT_COLUMN}",
        each
            let
                TemplateType = try Text.Upper(Text.Trim(Text.From([Template]))) otherwise "",
                MCValidationDate = try Date.From([#"Validation Date"]) otherwise null,
                BOAFSValidationDate = try Date.From([#"BO {BO_AFS_VALIDATION_DATE}"]) otherwise null,
                BOPINGNGValidationDate = try Date.From([#"BO {BO_PIN_GNG_VALIDATION_DATE}"]) otherwise null
            in
                if TemplateType = "OTHER" then MCValidationDate
                else if TemplateType = "AFS" then BOAFSValidationDate
                else if TemplateType = "GNG" or TemplateType = "PIN" then BOPINGNGValidationDate
                else null,
        type date
    ),
    TemplateCheckedBO = Table.AddColumn(
        BOValidationDate,
        "{BO_OUTPUT_COLUMN} Template Checked",
        each if (try Text.Upper(Text.Trim(Text.From([Template]))) otherwise "") = "OTHER"
            then null
            else [#"{BO_OUTPUT_COLUMN}"],
        type text
    ),
    BOServicePJ = Table.AddColumn(
        TemplateCheckedBO,
        "{BO_PJ_OUTPUT_COLUMN}",
        each if List.Contains({{"AFS", "GNG"}}, (try Text.Upper(Text.Trim(Text.From([Template]))) otherwise ""))
            then [#"BO {BO_PJ_COLUMN}"]
            else null,
        type text
    ),
    BOServiceRM = Table.AddColumn(
        BOServicePJ,
        "{BO_RM_OUTPUT_COLUMN}",
        each if List.Contains({{"AFS", "GNG"}}, (try Text.Upper(Text.Trim(Text.From([Template]))) otherwise ""))
            then [#"BO {BO_RM_COLUMN}"]
            else null,
        type text
    ),
    BOServiceJU = Table.AddColumn(
        BOServiceRM,
        "{BO_JU_OUTPUT_COLUMN}",
        each if List.Contains({{"AFS", "GNG"}}, (try Text.Upper(Text.Trim(Text.From([Template]))) otherwise ""))
            then [#"BO {BO_JU_COLUMN}"]
            else null,
        type text
    ),
    BOServiceECON = Table.AddColumn(
        BOServiceJU,
        "{BO_ECON_OUTPUT_COLUMN}",
        each if List.Contains({{"AFS", "GNG"}}, (try Text.Upper(Text.Trim(Text.From([Template]))) otherwise ""))
            then [#"BO {BO_ECON_COLUMN}"]
            else null,
        type text
    ),
    RemovedRawBO = Table.RemoveColumns(
        BOServiceECON,
        {m_list(BO_RAW_EXPANDED_COLUMNS)}
    ),
    MasterTable = Table.RenameColumns(
        RemovedRawBO,
        {{{{"{BO_OUTPUT_COLUMN} Template Checked", "{BO_OUTPUT_COLUMN}"}}}}
    ),
    OrderedMasterTable = Table.ReorderColumns(
        MasterTable,
        {m_list(MASTER_COLUMN_ORDER)},
        MissingField.Ignore
    )
in
    OrderedMasterTable"""


def build_master_dataframe() -> pd.DataFrame:
    mc = pd.read_excel(MC_DB, sheet_name="Database", header=1)
    mc = mc.dropna(how="all")
    mc = mc[MC_COLUMNS_B_TO_AC].copy()
    mc["Operation Number"] = pd.to_numeric(mc["Operation Number"], errors="coerce").astype("Int64")

    bo = pd.read_excel(BO_DB, sheet_name="BO Data", header=1)
    bo = bo.dropna(how="all")
    bo = bo[BO_SELECTED_COLUMNS].copy()
    bo["Operation"] = pd.to_numeric(bo["Operation"], errors="coerce").astype("Int64")
    bo = bo.drop_duplicates(subset=["Operation"])

    master = mc.merge(bo, how="left", left_on="Operation Number", right_on="Operation")
    master = master.drop(columns=["Operation"])
    master = master.rename(columns={BO_COLUMN_AB: BO_OUTPUT_COLUMN})
    template_type = master["Template"].astype(str).str.strip().str.upper()
    master[BO_VALIDATION_OUTPUT_COLUMN] = pd.NA
    master.loc[template_type.eq("AFS"), BO_VALIDATION_OUTPUT_COLUMN] = master.loc[
        template_type.eq("AFS"), BO_AFS_VALIDATION_DATE
    ]
    master.loc[template_type.isin(["GNG", "PIN"]), BO_VALIDATION_OUTPUT_COLUMN] = master.loc[
        template_type.isin(["GNG", "PIN"]), BO_PIN_GNG_VALIDATION_DATE
    ]
    master.loc[template_type.eq("OTHER"), BO_VALIDATION_OUTPUT_COLUMN] = master.loc[
        template_type.eq("OTHER"), "Validation Date"
    ]

    afs_gng_mask = template_type.isin(["AFS", "GNG"])
    for source_column, output_column in BO_SERVICE_COLUMN_MAP.items():
        master[output_column] = pd.NA
        master.loc[afs_gng_mask, output_column] = master.loc[afs_gng_mask, source_column]

    master = master.drop(
        columns=[BO_PIN_GNG_VALIDATION_DATE, BO_AFS_VALIDATION_DATE, *BO_SERVICE_COLUMN_MAP.keys()]
    )

    other_mask = template_type.eq("OTHER")
    master.loc[other_mask, BO_OUTPUT_COLUMN] = pd.NA
    master = master[[column for column in MASTER_COLUMN_ORDER if column in master.columns]]
    return master


def excel_value(value):
    if pd.isna(value):
        return None
    if hasattr(value, "to_pydatetime"):
        return value.to_pydatetime()
    return value


def write_workbook(master: pd.DataFrame) -> None:
    if MASTER.exists():
        MASTER.unlink()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Master Table"

    header_fill = PatternFill("solid", fgColor="0F6CBD")
    header_font = Font(color="FFFFFF", bold=True)

    sheet.append(list(master.columns))
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in master.itertuples(index=False, name=None):
        sheet.append([excel_value(value) for value in row])

    last_col = get_column_letter(master.shape[1])
    last_row = master.shape[0] + 1
    table = Table(displayName="Master_Table", ref=f"A1:{last_col}{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{last_col}{last_row}"

    for idx, col_name in enumerate(master.columns, start=1):
        width = min(max(len(str(col_name)) + 2, 12), 48)
        sheet.column_dimensions[get_column_letter(idx)].width = width

    readme = workbook.create_sheet("README")
    readme["A1"] = "Master Table"
    readme["A1"].font = Font(bold=True, size=14)
    readme["A3"] = "Loaded table: MC_Note_Datebase.xlsx / Database / columns B:AC."
    readme["A4"] = f"Added BO column AB as: {BO_OUTPUT_COLUMN}."
    readme["A5"] = "Join key: MC Operation Number = BO Operation."
    readme["A6"] = "For AFS and GNG only, added BO PJ/RM/JU/ECON service fields at the end of the table; BO ECON uses the BO SG division field."
    readme["A7"] = "Power Query M code is included in the 'Power Query M' sheet and saved beside this workbook."
    readme["A8"] = "Note: Excel COM was unstable in this environment when embedding the query connection, so the table is preloaded from the same logic."
    readme.column_dimensions["A"].width = 120

    query_sheet = workbook.create_sheet("Power Query M")
    query_sheet["A1"] = "Paste this into Power Query Advanced Editor if you want the workbook-native refreshable query:"
    query_sheet["A1"].font = Font(bold=True)
    for idx, line in enumerate(build_power_query_formula().splitlines(), start=3):
        query_sheet.cell(row=idx, column=1, value=line)
    query_sheet.column_dimensions["A"].width = 160

    workbook.save(MASTER)


def main() -> None:
    if not MC_DB.exists():
        raise FileNotFoundError(MC_DB)
    if not BO_DB.exists():
        raise FileNotFoundError(BO_DB)

    master = build_master_dataframe()
    write_workbook(master)
    QUERY_FILE.write_text(build_power_query_formula(), encoding="utf-8")

    matched = int(master[BO_OUTPUT_COLUMN].notna().sum())
    print(f"Created {MASTER}")
    print(f"Rows: {len(master):,}; columns: {master.shape[1]:,}; BO AB matches: {matched:,}")
    print(f"Wrote {QUERY_FILE}")


if __name__ == "__main__":
    main()
