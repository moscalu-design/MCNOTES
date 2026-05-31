import csv
import shutil
import sys
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
OTHER_DIR = ROOT / "OTHER_Package" / "OTHER_File_Folder"
DATABASE_WORKBOOK = ROOT / "Data Analysis" / "MC_Note_Datebase.xlsx"
OUTPUT_WORKBOOK = ROOT / "Data Analysis" / "Workbooks" / "MC_Note_Datebase_other_date_fix.xlsx"
AUDIT_CSV = ROOT / "Data Analysis" / "Exports" / "other_validation_date_audit.csv"

HEADER_ROW = 2
DATABASE_SHEET = "Database"


sys.path.insert(0, str(ROOT))
from OTHER_Package.other_extractor import (  # noqa: E402
    extract_doc_date_and_author_top_right,
    extract_mc_metadata,
    normalize_validation_date,
    repair_mojibake_text,
)


def year_from_path(path: Path) -> str:
    for part in path.relative_to(OTHER_DIR).parts:
        if part.isdigit() and len(part) == 4:
            return part
    return ""


def as_ddmmyyyy(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return normalize_validation_date(text)


def filename_match_key(value: str) -> str:
    text = repair_mojibake_text(value)
    text = " ".join(text.replace("\u00a0", " ").split())
    return text.casefold()


def date_to_excel(value: str):
    value = as_ddmmyyyy(value)
    if not value:
        return ""
    try:
        return datetime.strptime(value, "%d/%m/%Y")
    except ValueError:
        return value


def extract_pdf_date(pdf_path: Path) -> dict:
    try:
        visible_date, _author, debug = extract_doc_date_and_author_top_right(pdf_path)
        _mc_type, filename_date = extract_mc_metadata(pdf_path.name)
        chosen_date = normalize_validation_date(visible_date or filename_date)
        status = "visible_date" if visible_date else "filename_fallback" if filename_date else "missing_pdf_date"
        error = ""
    except Exception as exc:  # Keep the audit moving; failures are reported row-by-row.
        visible_date = ""
        filename_date = ""
        chosen_date = ""
        debug = {}
        status = "pdf_error"
        error = str(exc)

    return {
        "pdf_path": str(pdf_path),
        "pdf_file_name": pdf_path.name,
        "pdf_year": year_from_path(pdf_path),
        "visible_header_date": normalize_validation_date(visible_date),
        "filename_date": normalize_validation_date(filename_date),
        "new_validation_date": chosen_date,
        "pdf_date_status": status,
        "date_method": debug.get("date_method", ""),
        "date_trigger_text": debug.get("date_trigger_text", ""),
        "pdf_error": error,
    }


def build_pdf_date_lookup():
    pdfs_by_name = defaultdict(list)
    for pdf_path in sorted(OTHER_DIR.rglob("*.pdf")):
        pdfs_by_name[pdf_path.name].append(pdf_path)

    lookup = {}
    normalized_lookup = {}
    normalized_duplicates = set()
    duplicate_names = 0
    conflicting_duplicate_names = 0
    year_counts = Counter()

    for file_name, paths in sorted(pdfs_by_name.items()):
        extracted = [extract_pdf_date(path) for path in paths]
        for item in extracted:
            year_counts[item["pdf_year"]] += 1

        unique_dates = {item["new_validation_date"] for item in extracted if item["new_validation_date"]}
        if len(paths) > 1:
            duplicate_names += 1
        if len(unique_dates) > 1:
            conflicting_duplicate_names += 1
            chosen = extracted[0]
            chosen = {**chosen, "pdf_date_status": "ambiguous_duplicate_filename"}
        else:
            chosen = extracted[0]

        chosen["pdf_match_count"] = len(paths)
        chosen["duplicate_pdf_paths"] = " | ".join(item["pdf_path"] for item in extracted)
        lookup[file_name] = chosen
        key = filename_match_key(file_name)
        if key in normalized_lookup:
            normalized_duplicates.add(key)
        else:
            normalized_lookup[key] = chosen

    for key in normalized_duplicates:
        normalized_lookup.pop(key, None)

    return lookup, normalized_lookup, year_counts, duplicate_names, conflicting_duplicate_names, len(normalized_duplicates)


def header_map(ws):
    return {ws.cell(HEADER_ROW, col).value: col for col in range(1, ws.max_column + 1)}


def ensure_output_columns(ws, validation_col: int):
    headers = header_map(ws)
    fix_col = headers.get("Other date fix? (Y/N)")
    new_date_col = headers.get("New Validation Date")
    if fix_col and new_date_col:
        return fix_col, new_date_col

    ws.insert_cols(validation_col + 1, 2)
    fix_col = validation_col + 1
    new_date_col = validation_col + 2
    ws.cell(HEADER_ROW, fix_col).value = "Other date fix? (Y/N)"
    ws.cell(HEADER_ROW, new_date_col).value = "New Validation Date"
    for col in (fix_col, new_date_col):
        ws.cell(HEADER_ROW, col)._style = ws.cell(HEADER_ROW, validation_col)._style
    return fix_col, new_date_col


def update_workbook(pdf_lookup: dict, normalized_pdf_lookup: dict):
    wb = load_workbook(DATABASE_WORKBOOK)
    ws = wb[DATABASE_SHEET]
    headers = header_map(ws)
    template_col = headers["Template"]
    source_col = headers["Source"]
    file_col = headers["File Name"]
    validation_col = headers["Validation Date"]
    fix_col, new_date_col = ensure_output_columns(ws, validation_col)

    audit_rows = []
    status_counts = Counter()
    other_rows = 0

    for row_idx in range(HEADER_ROW + 1, ws.max_row + 1):
        template = ws.cell(row_idx, template_col).value
        if template != "Other":
            ws.cell(row_idx, fix_col).value = ""
            ws.cell(row_idx, new_date_col).value = ""
            continue

        other_rows += 1
        file_name = str(ws.cell(row_idx, file_col).value or "").strip()
        db_date = as_ddmmyyyy(ws.cell(row_idx, validation_col).value)
        pdf_info = pdf_lookup.get(file_name)
        filename_match_method = "exact_filename" if pdf_info else ""
        if not pdf_info:
            pdf_info = normalized_pdf_lookup.get(filename_match_key(file_name))
            filename_match_method = "normalized_filename" if pdf_info else ""

        if not pdf_info:
            fix_flag = ""
            new_date = ""
            status = "no_matching_pdf_file_name"
        else:
            new_date = pdf_info["new_validation_date"]
            if not new_date:
                fix_flag = ""
                status = pdf_info["pdf_date_status"]
            elif db_date == new_date:
                fix_flag = "N"
                status = "date_matches"
            else:
                fix_flag = "Y"
                status = "date_mismatch"

        ws.cell(row_idx, fix_col).value = fix_flag
        if new_date:
            ws.cell(row_idx, new_date_col).value = date_to_excel(new_date)
            ws.cell(row_idx, new_date_col).number_format = "DD/MM/YYYY"
        else:
            ws.cell(row_idx, new_date_col).value = ""

        status_counts[status] += 1
        audit_rows.append({
            "db_excel_row": row_idx,
            "Source": ws.cell(row_idx, source_col).value,
            "Template": template,
            "File Name": file_name,
            "Existing Validation Date": db_date,
            "Other date fix? (Y/N)": fix_flag,
            "New Validation Date": new_date,
            "Status": status,
            "Filename Match Method": filename_match_method,
            **(pdf_info or {
                "pdf_path": "",
                "pdf_file_name": "",
                "pdf_year": "",
                "visible_header_date": "",
                "filename_date": "",
                "pdf_date_status": "",
                "date_method": "",
                "date_trigger_text": "",
                "pdf_error": "",
                "pdf_match_count": 0,
                "duplicate_pdf_paths": "",
            }),
        })

    OUTPUT_WORKBOOK.parent.mkdir(parents=True, exist_ok=True)
    AUDIT_CSV.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_WORKBOOK)
    wb.close()

    fieldnames = list(audit_rows[0].keys()) if audit_rows else []
    with AUDIT_CSV.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(audit_rows)

    return other_rows, status_counts


def main():
    if not DATABASE_WORKBOOK.exists():
        raise SystemExit(f"Database workbook not found: {DATABASE_WORKBOOK}")

    shutil.copy2(DATABASE_WORKBOOK, OUTPUT_WORKBOOK.with_suffix(".source-backup.xlsx"))
    (
        pdf_lookup,
        normalized_pdf_lookup,
        year_counts,
        duplicate_names,
        conflicting_duplicate_names,
        normalized_duplicate_names,
    ) = build_pdf_date_lookup()
    other_rows, status_counts = update_workbook(pdf_lookup, normalized_pdf_lookup)

    print("OTHER validation date audit complete")
    print(f"PDFs scanned: {sum(year_counts.values())}")
    for year, count in sorted(year_counts.items()):
        print(f"  {year or 'unknown'}: {count}")
    print(f"Unique PDF file names: {len(pdf_lookup)}")
    print(f"Duplicate file names: {duplicate_names}")
    print(f"Conflicting duplicate file names: {conflicting_duplicate_names}")
    print(f"Ambiguous normalized file names: {normalized_duplicate_names}")
    print(f"Database OTHER rows: {other_rows}")
    for status, count in sorted(status_counts.items()):
        print(f"  {status}: {count}")
    print(f"Updated workbook: {OUTPUT_WORKBOOK}")
    print(f"Audit CSV: {AUDIT_CSV}")


if __name__ == "__main__":
    main()
