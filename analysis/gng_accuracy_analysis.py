import csv
import html
import importlib.util
import json
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean

import fitz
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
GNG_DIR = ROOT / "GNG Folder"
OUTPUT_DIR = ROOT / "outputs" / "gng_accuracy"
LIMIT = 25

CSV_PREFIXES = [
    "OPS", "GLO", "PJ", "RM", "OCCO", "JU", "ECON",
    "CFC", "EIF", "FI", "IG", "PMM", "SG", "GIS", "HR", "OTHER"
]

BASE_COLUMNS = [
    "Template",
    "Extraction",
    "MC_Note_Type",
    "File Name",
    "Operation Number",
    "Validation Date",
    "Author",
    "Document Page Count",
    "Page count before opinion",
    "Annex Page Count",
    "Text Before Opinions"
]

FIELDS = BASE_COLUMNS + CSV_PREFIXES

WORD_RE = re.compile(r"[^\W_]+(?:'[^\W_]+)?", re.UNICODE)
DATE_RE = re.compile(r"\b(\d{2})(?:/|\.)(\d{2})(?:/|\.)(\d{4})\b")
OP_RE = re.compile(r"\b(?:20\d{2}[- ]?\d{4}|20\d{6})\b")
SECTION6_RE = re.compile(r"\b6\.\s*Services['’]?\s+acknowledgement", re.IGNORECASE)
EXPECTED_TIMETABLE_RE = re.compile(r"\bExpected\s+timetable\b", re.IGNORECASE)
GNG_VALIDATION_RE = re.compile(r"\bGNG\s+validation\b", re.IGNORECASE)

PREFIX_HEAD_RE = re.compile(
    r"^\s*(GR\s*&\s*C\s*[-/\u2010-\u2015\u2212]?\s*(?:RM|OCCO)|OCCO|ECON|PJ|JU)\b",
    re.IGNORECASE
)

BOILERPLATE_RE = re.compile(
    r"\bagrees\s+to\s+commence\s+appraisal\s+of\s+the\s+project\b",
    re.IGNORECASE
)

NOISE_LINE_RE = re.compile(
    r"^\s*(?:Corporate Use|-\s*\d+\s*-|\d+\s+of\s+\d+|MC[_\s-]?GNG.*)\s*$",
    re.IGNORECASE
)


def load_gng_module():
    spec = importlib.util.spec_from_file_location("gng_extractor", ROOT / "gng_extractor.py")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def normalize_text(value):
    value = str(value or "")
    value = value.replace("’", "'").replace("‘", "'")
    value = value.replace("\u00ad", "")
    value = value.replace("\u00a0", " ")
    value = re.sub(r"[\u2010-\u2015\u2212]", "-", value)
    return re.sub(r"\s+", " ", value).strip()


def normalize_compare(value):
    value = normalize_text(value)
    if value.lower() == "none":
        return ""
    return value


def normalize_date(value):
    match = DATE_RE.search(str(value or ""))
    if not match:
        return ""
    return f"{match.group(1)}/{match.group(2)}/{match.group(3)}"


def canonical_prefix(value):
    text = normalize_text(value).upper()
    text = re.sub(r"\s+", "", text)
    text = re.sub(r"[-/]+", "-", text)
    if re.match(r"^GR&C-?RM\b", text):
        return "GR&C-RM"
    if re.match(r"^GR&C-?OCCO\b", text):
        return "GR&C-OCCO"
    return text


def export_prefix(value):
    if value == "GR&C-RM":
        return "RM"
    if value in {"GR&C-OCCO", "OCCO"}:
        return "OCCO"
    return value


def count_words(value):
    text = normalize_text(value)
    text = re.sub(r"(?m)^\s*[oO]\s+", "", text)
    text = text.replace("&", "").replace("-", "").replace("/", "")
    text = re.sub(r"\s+", " ", text)
    return len(WORD_RE.findall(text))


def clean_file_title(value):
    return str(value).replace(",", "")


def read_page_lines(doc):
    all_lines = []
    order = 0
    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        grouped = {}
        for item in page.get_text("words"):
            x0, y0, x1, y1, word, block, line, word_no = item
            grouped.setdefault((block, line), []).append((x0, y0, x1, y1, word))

        page_lines = []
        for parts in grouped.values():
            parts = sorted(parts, key=lambda part: (part[0], part[1]))
            text = " ".join(part[4] for part in parts)
            page_lines.append({
                "page_index": page_index,
                "page": page_index + 1,
                "x0": min(part[0] for part in parts),
                "y0": min(part[1] for part in parts),
                "x1": max(part[2] for part in parts),
                "y1": max(part[3] for part in parts),
                "text": normalize_text(text),
            })

        page_lines.sort(key=lambda row: (row["y0"], row["x0"]))
        for line in page_lines:
            line["order"] = order
            order += 1
            all_lines.append(line)
    return all_lines


def line_is_noise(line):
    text = normalize_text(line["text"])
    return not text or bool(NOISE_LINE_RE.match(text))


def find_section6_line(lines):
    for line in lines:
        if SECTION6_RE.search(line["text"]):
            return line
    return None


def find_expected_timetable_line(lines, section_line):
    start_order = section_line["order"] if section_line else -1
    for line in lines:
        if line["order"] <= start_order:
            continue
        if EXPECTED_TIMETABLE_RE.search(line["text"]):
            return line
    return None


def extract_visible_operation_number(file_name, lines):
    title_candidates = []
    for line in lines[:120]:
        text = line["text"]
        if "GNG" in text.upper() or "MCGNG" in text.upper():
            title_candidates.append(text)
    searchable = "\n".join(title_candidates + [file_name])
    match = OP_RE.search(searchable)
    if match:
        return match.group(0).replace("-", "").replace(" ", "")

    match = OP_RE.search("\n".join(line["text"] for line in lines[:200]))
    if match:
        return match.group(0).replace("-", "").replace(" ", "")
    return ""


def extract_visible_validation_date(lines):
    for line in lines[:120]:
        if "luxembourg" not in line["text"].lower():
            continue
        date = normalize_date(line["text"])
        if date:
            return date
    for line in lines[:120]:
        date = normalize_date(line["text"])
        if date:
            return date
    return ""


def extract_visible_author(lines):
    first_page_lines = [line for line in lines if line["page"] == 1]
    date_index = None
    author_line_re = re.compile(r"^[A-Z]{2}[A-Za-z0-9&.\- ]*(?:/[A-Za-z0-9&().,'\- ]+)+$")

    for idx, line in enumerate(first_page_lines):
        if normalize_date(line["text"]):
            date_index = idx
            break

    if date_index is not None:
        for candidate in first_page_lines[date_index + 1:date_index + 16]:
            candidate_text = normalize_text(candidate["text"])
            if author_line_re.fullmatch(candidate_text):
                return candidate_text

    for idx, line in enumerate(first_page_lines):
        text = line["text"]
        if "loan officer" not in text.lower() and not text.startswith("OPS Directorate"):
            continue
        for candidate in first_page_lines[idx + 1:idx + 6]:
            candidate_text = normalize_text(candidate["text"])
            if candidate_text.startswith("OPS/"):
                return candidate_text
    return ""


def is_prefix_heading(text):
    normalized = normalize_text(text)
    if BOILERPLATE_RE.search(normalized):
        return False
    match = PREFIX_HEAD_RE.match(normalized)
    if not match:
        return False
    word_count = count_words(normalized)
    return word_count <= 6


def extract_visible_prefix_counts(lines, section_line, timetable_line):
    if not section_line:
        return {}, {"status": "Section 6 not found"}

    end_order = timetable_line["order"] if timetable_line else 10**9
    section_lines = [
        line for line in lines
        if section_line["order"] < line["order"] < end_order
        and not line_is_noise(line)
    ]

    counts = Counter()
    current_prefix = ""
    segment_lines = []
    segments = []

    def flush():
        nonlocal segment_lines, current_prefix
        if current_prefix and segment_lines:
            words = sum(count_words(line["text"]) for line in segment_lines)
            counts[export_prefix(current_prefix)] += words
            segments.append({
                "prefix": current_prefix,
                "export_prefix": export_prefix(current_prefix),
                "words": words,
                "line_count": len(segment_lines),
                "preview": " | ".join(line["text"] for line in segment_lines[:4])
            })
        segment_lines = []

    for line in section_lines:
        text = line["text"]
        if is_prefix_heading(text):
            detected = canonical_prefix(PREFIX_HEAD_RE.match(text).group(1))
            exported = export_prefix(detected)
            if exported in CSV_PREFIXES and detected != current_prefix:
                flush()
                current_prefix = detected
        if current_prefix:
            segment_lines.append(line)

    flush()
    return dict(counts), {"status": "OK", "segments": segments}


def extract_visible_text_before_opinions(lines, section_line):
    if not section_line:
        return "check"
    total = 0
    for line in lines:
        if line["order"] >= section_line["order"]:
            break
        if line_is_noise(line):
            continue
        total += count_words(line["text"])
    return total


def extract_visible_annex_count(doc, lines, timetable_line):
    if not timetable_line:
        return "check"
    for line in lines:
        if line["order"] <= timetable_line["order"]:
            continue
        if GNG_VALIDATION_RE.search(line["text"]):
            return max(doc.page_count - line["page"], 0)
    return "check"


def build_actual_row(pdf_path):
    with fitz.open(pdf_path) as doc:
        lines = read_page_lines(doc)
        section_line = find_section6_line(lines)
        timetable_line = find_expected_timetable_line(lines, section_line)
        prefix_counts, prefix_debug = extract_visible_prefix_counts(lines, section_line, timetable_line)

        row = {
            "Template": "GNG",
            "Extraction": "Automated",
            "MC_Note_Type": "NOTEMCDEC",
            "File Name": clean_file_title(pdf_path.name),
            "Operation Number": extract_visible_operation_number(pdf_path.name, lines),
            "Validation Date": extract_visible_validation_date(lines),
            "Author": extract_visible_author(lines),
            "Document Page Count": doc.page_count,
            "Page count before opinion": "check" if not section_line else section_line["page"],
            "Annex Page Count": extract_visible_annex_count(doc, lines, timetable_line),
            "Text Before Opinions": extract_visible_text_before_opinions(lines, section_line),
        }

        for prefix in CSV_PREFIXES:
            row[prefix] = prefix_counts.get(prefix, "")

        return row, {
            "section_found": bool(section_line),
            "section_page": "" if not section_line else section_line["page"],
            "expected_timetable_found": bool(timetable_line),
            "expected_timetable_page": "" if not timetable_line else timetable_line["page"],
            "prefix_debug": prefix_debug,
        }


def coerce_number(value):
    text = normalize_compare(value)
    if text == "":
        return None
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        number = float(text)
        return int(number) if number.is_integer() else number
    return None


def compare_values(field, script_value, actual_value):
    script_norm = normalize_compare(script_value)
    actual_norm = normalize_compare(actual_value)
    script_num = coerce_number(script_norm)
    actual_num = coerce_number(actual_norm)
    delta = ""
    pct_delta = ""

    if field in {"Validation Date"}:
        script_norm = normalize_date(script_norm)
        actual_norm = normalize_date(actual_norm)

    if script_num is not None and actual_num is not None:
        delta_value = script_num - actual_num
        delta = delta_value
        if actual_num not in (0, None):
            pct_delta = delta_value / actual_num
        match = delta_value == 0
    else:
        match = script_norm == actual_norm

    if actual_norm == "" and script_norm == "":
        status = "both blank"
    elif match:
        status = "match"
    else:
        status = "mismatch"

    return {
        "script_normalized": script_norm,
        "actual_normalized": actual_norm,
        "match": match,
        "status": status,
        "delta": delta,
        "pct_delta": pct_delta,
    }


def build_script_row(gng, pdf_path):
    result = gng.process_pdf(pdf_path)
    return gng.build_csv_row(result)


def write_csv(path, rows, fieldnames):
    with open(path, "w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def calculate_stats(comparison_rows, document_rows):
    field_stats = []
    by_field = defaultdict(list)
    for row in comparison_rows:
        by_field[row["Field"]].append(row)

    for field in FIELDS:
        rows = by_field[field]
        compared = len(rows)
        matches = sum(1 for row in rows if row["Match"] == "YES")
        mismatches = compared - matches
        numeric_deltas = [
            abs(float(row["Numeric Delta"]))
            for row in rows
            if str(row["Numeric Delta"]).strip() not in {"", "None"}
        ]
        field_stats.append({
            "Field": field,
            "Compared": compared,
            "Matches": matches,
            "Mismatches": mismatches,
            "Match Rate": matches / compared if compared else "",
            "Mean Abs Delta": mean(numeric_deltas) if numeric_deltas else "",
            "Max Abs Delta": max(numeric_deltas) if numeric_deltas else "",
        })

    overall_compared = len(comparison_rows)
    overall_matches = sum(1 for row in comparison_rows if row["Match"] == "YES")
    overall = {
        "Documents": len(document_rows),
        "Fields Compared": overall_compared,
        "Field Matches": overall_matches,
        "Field Mismatches": overall_compared - overall_matches,
        "Overall Match Rate": overall_matches / overall_compared if overall_compared else 0,
        "Perfect Documents": sum(1 for row in document_rows if row["Mismatched Fields"] == 0),
    }

    return overall, field_stats


def autosize_sheet(sheet):
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 10), 55)


def write_workbook(path, overall, field_stats, document_rows, comparison_rows, script_rows, actual_rows):
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    good_fill = PatternFill("solid", fgColor="DCFCE7")
    bad_fill = PatternFill("solid", fgColor="FEE2E2")
    warn_fill = PatternFill("solid", fgColor="FEF3C7")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def add_sheet(name, rows):
        sheet = workbook.create_sheet(name)
        if not rows:
            return sheet
        headers = list(rows[0].keys())
        sheet.append(headers)
        for row in rows:
            sheet.append([row.get(header, "") for header in headers])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        autosize_sheet(sheet)
        return sheet

    summary_sheet = workbook.create_sheet("Summary")
    summary_rows = [
        ["Metric", "Value"],
        ["Documents analysed", overall["Documents"]],
        ["Fields compared", overall["Fields Compared"]],
        ["Field matches", overall["Field Matches"]],
        ["Field mismatches", overall["Field Mismatches"]],
        ["Overall match rate", overall["Overall Match Rate"]],
        ["Perfect documents", overall["Perfect Documents"]],
        ["Counting basis", "Full service block count"],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    autosize_sheet(summary_sheet)

    field_sheet = add_sheet("Field Stats", field_stats)
    doc_sheet = add_sheet("Document Summary", document_rows)
    comparison_sheet = add_sheet("Field Comparison", comparison_rows)
    add_sheet("Script Output", script_rows)
    add_sheet("Visible Actuals", actual_rows)

    for sheet in [field_sheet, doc_sheet, comparison_sheet]:
        for row in sheet.iter_rows(min_row=2):
            values = [cell.value for cell in row]
            if "YES" in values or "Match Rate" in [cell.value for cell in sheet[1]]:
                pass
            for cell in row:
                if cell.value == "YES" or cell.value == 1:
                    cell.fill = good_fill
                elif cell.value == "NO" or cell.value == "mismatch":
                    cell.fill = bad_fill
                elif cell.value == "both blank":
                    cell.fill = warn_fill

    workbook.save(path)


def percent(value):
    if value == "":
        return ""
    return f"{float(value) * 100:.1f}%"


def write_html(path, overall, field_stats, document_rows, comparison_rows, selected_files):
    worst_fields = sorted(field_stats, key=lambda row: (row["Match Rate"] if row["Match Rate"] != "" else 0))[:8]
    worst_docs = sorted(document_rows, key=lambda row: row["Match Rate"])[:8]
    mismatches = [row for row in comparison_rows if row["Match"] == "NO"]

    def table(rows, columns):
        cells = ["<table><thead><tr>"]
        for column in columns:
            cells.append(f"<th>{html.escape(column)}</th>")
        cells.append("</tr></thead><tbody>")
        for row in rows:
            cells.append("<tr>")
            for column in columns:
                value = row.get(column, "")
                if isinstance(value, float) and "Rate" in column:
                    value = percent(value)
                cells.append(f"<td>{html.escape(str(value))}</td>")
            cells.append("</tr>")
        cells.append("</tbody></table>")
        return "".join(cells)

    html_text = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <title>GNG Accuracy Analysis</title>
  <style>
    :root {{
      --ink: #172033;
      --muted: #64748b;
      --line: #dbe3ef;
      --panel: #ffffff;
      --bg: #f4f7fb;
      --green: #15803d;
      --red: #b91c1c;
      --amber: #a16207;
      --blue: #1d4ed8;
    }}
    body {{
      margin: 0;
      font-family: Aptos, Segoe UI, Arial, sans-serif;
      background: var(--bg);
      color: var(--ink);
    }}
    header {{
      background: #0f172a;
      color: white;
      padding: 28px 34px;
    }}
    header h1 {{
      margin: 0 0 8px;
      font-size: 28px;
      letter-spacing: 0;
    }}
    header p {{
      margin: 0;
      color: #cbd5e1;
    }}
    main {{
      padding: 24px 34px 42px;
      max-width: 1320px;
      margin: 0 auto;
    }}
    .grid {{
      display: grid;
      grid-template-columns: repeat(4, minmax(180px, 1fr));
      gap: 14px;
      margin-bottom: 18px;
    }}
    .metric {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 16px;
    }}
    .metric span {{
      display: block;
      color: var(--muted);
      font-size: 13px;
      margin-bottom: 8px;
    }}
    .metric strong {{
      font-size: 28px;
    }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 18px;
      margin-top: 16px;
    }}
    h2 {{
      margin: 0 0 12px;
      font-size: 18px;
    }}
    table {{
      border-collapse: collapse;
      width: 100%;
      font-size: 13px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 8px 10px;
      text-align: left;
      vertical-align: top;
    }}
    th {{
      background: #eef2f7;
      color: #263244;
      font-weight: 700;
    }}
    .note {{
      color: var(--muted);
      line-height: 1.45;
      margin-top: 8px;
    }}
    .good {{ color: var(--green); }}
    .bad {{ color: var(--red); }}
    .amber {{ color: var(--amber); }}
    .files {{
      columns: 2;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.45;
    }}
  </style>
</head>
<body>
  <header>
    <h1>GNG Script Accuracy Analysis</h1>
    <p>25-document sample, comparing script output against independently extracted visible PDF values. Counting basis: full service block count.</p>
  </header>
  <main>
    <div class="grid">
      <div class="metric"><span>Documents Analysed</span><strong>{overall["Documents"]}</strong></div>
      <div class="metric"><span>Overall Field Match Rate</span><strong>{percent(overall["Overall Match Rate"])}</strong></div>
      <div class="metric"><span>Field Mismatches</span><strong>{overall["Field Mismatches"]}</strong></div>
      <div class="metric"><span>Perfect Documents</span><strong>{overall["Perfect Documents"]}</strong></div>
    </div>

    <section>
      <h2>Executive Summary</h2>
      <p class="note">The independent check reads the PDF through visible word positions and line geometry. It then compares those values to the script output field-by-field. Empty prefix columns are treated as matches when both script and visible extraction are blank.</p>
      <p class="note">The most important review areas are fields with low match rate or large numeric deltas. See the detailed CSV/workbook for every document-field comparison.</p>
    </section>

    <section>
      <h2>Lowest Field Match Rates</h2>
      {table(worst_fields, ["Field", "Compared", "Matches", "Mismatches", "Match Rate", "Mean Abs Delta", "Max Abs Delta"])}
    </section>

    <section>
      <h2>Documents With Most Differences</h2>
      {table(worst_docs, ["File Name", "Matched Fields", "Compared Fields", "Mismatched Fields", "Match Rate", "Mismatch Fields"])}
    </section>

    <section>
      <h2>Mismatch Detail</h2>
      {table(mismatches[:80], ["File Name", "Field", "Script Value", "Actual Value", "Numeric Delta", "Status"])}
    </section>

    <section>
      <h2>Selected Files</h2>
      <div class="files">{"<br>".join(html.escape(path.name) for path in selected_files)}</div>
    </section>
  </main>
</body>
</html>
"""
    path.write_text(html_text, encoding="utf-8")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    gng = load_gng_module()
    pdf_files = sorted(GNG_DIR.rglob("*.pdf"))[:LIMIT]

    script_rows = []
    actual_rows = []
    debug_rows = []
    comparison_rows = []
    document_rows = []

    for index, pdf_path in enumerate(pdf_files, 1):
        script_row = build_script_row(gng, pdf_path)
        actual_row, debug = build_actual_row(pdf_path)
        script_row = {field: script_row.get(field, "") for field in FIELDS}
        actual_row = {field: actual_row.get(field, "") for field in FIELDS}

        script_rows.append({"Sample #": index, **script_row})
        actual_rows.append({"Sample #": index, **actual_row})
        debug_rows.append({
            "Sample #": index,
            "File Name": pdf_path.name,
            "Section Found": "YES" if debug["section_found"] else "NO",
            "Section Page": debug["section_page"],
            "Expected Timetable Found": "YES" if debug["expected_timetable_found"] else "NO",
            "Expected Timetable Page": debug["expected_timetable_page"],
            "Prefix Segment Count": len(debug["prefix_debug"].get("segments", [])),
            "Prefix Debug": json.dumps(debug["prefix_debug"], ensure_ascii=False),
        })

        mismatch_fields = []
        compared_fields = 0
        matched_fields = 0

        for field in FIELDS:
            comparison = compare_values(field, script_row.get(field, ""), actual_row.get(field, ""))
            compared_fields += 1
            if comparison["match"]:
                matched_fields += 1
            else:
                mismatch_fields.append(field)

            comparison_rows.append({
                "Sample #": index,
                "File Name": pdf_path.name,
                "Field": field,
                "Script Value": script_row.get(field, ""),
                "Actual Value": actual_row.get(field, ""),
                "Script Normalized": comparison["script_normalized"],
                "Actual Normalized": comparison["actual_normalized"],
                "Match": "YES" if comparison["match"] else "NO",
                "Numeric Delta": comparison["delta"],
                "Percent Delta": comparison["pct_delta"],
                "Status": comparison["status"],
            })

        document_rows.append({
            "Sample #": index,
            "File Name": pdf_path.name,
            "Matched Fields": matched_fields,
            "Compared Fields": compared_fields,
            "Mismatched Fields": len(mismatch_fields),
            "Match Rate": matched_fields / compared_fields if compared_fields else 0,
            "Mismatch Fields": ", ".join(mismatch_fields),
        })

    overall, field_stats = calculate_stats(comparison_rows, document_rows)

    comparison_csv = OUTPUT_DIR / "gng_accuracy_field_comparison.csv"
    document_csv = OUTPUT_DIR / "gng_accuracy_document_summary.csv"
    field_stats_csv = OUTPUT_DIR / "gng_accuracy_field_stats.csv"
    script_csv = OUTPUT_DIR / "gng_script_output_25.csv"
    actual_csv = OUTPUT_DIR / "gng_visible_actuals_25.csv"
    debug_csv = OUTPUT_DIR / "gng_visible_extraction_debug.csv"
    workbook_path = OUTPUT_DIR / "gng_accuracy_analysis.xlsx"
    html_path = OUTPUT_DIR / "index.html"

    write_csv(comparison_csv, comparison_rows, list(comparison_rows[0].keys()))
    write_csv(document_csv, document_rows, list(document_rows[0].keys()))
    write_csv(field_stats_csv, field_stats, list(field_stats[0].keys()))
    write_csv(script_csv, script_rows, list(script_rows[0].keys()))
    write_csv(actual_csv, actual_rows, list(actual_rows[0].keys()))
    write_csv(debug_csv, debug_rows, list(debug_rows[0].keys()))
    write_workbook(workbook_path, overall, field_stats, document_rows, comparison_rows, script_rows, actual_rows)
    write_html(html_path, overall, field_stats, document_rows, comparison_rows, pdf_files)

    print("GNG accuracy analysis complete")
    print(f"Documents analysed: {overall['Documents']}")
    print(f"Overall match rate: {overall['Overall Match Rate']:.3%}")
    print(f"Field mismatches: {overall['Field Mismatches']}")
    print(f"Comparison CSV: {comparison_csv}")
    print(f"Workbook: {workbook_path}")
    print(f"HTML summary: {html_path}")


if __name__ == "__main__":
    main()
