#1.0 27/05/2026
import argparse
import csv
import math
import re
import statistics
import unicodedata
from collections import defaultdict
from pathlib import Path

import fitz

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


DEBUG = False
STATUS_OUTPUT = DEBUG
DEBUG_TEXT_CHAR_LIMIT = None

FOLDER_PATH = Path(__file__).resolve().parent / "OTHER_File_Folder"
OUTPUT_DIR = Path(__file__).resolve().parent / "outputs"
PREFIX_MERGE_DIR = Path(__file__).resolve().parent
PREFIX_MERGE_BASENAME = "PREFIX_MERGE"

NA_VALUE = "N/A"
TEMPLATE_VALUE = "Other"
EXTRACTION_VALUE = "Automated"

OUTPUT_PREFIXES = [
    "OPS", "GLO", "PJ", "RM", "OCCO", "JU", "ECON", "CFC",
    "EIF", "FI", "IG", "PMM", "SG", "GIS", "HR", "OTHER",
]

BASE_COLUMNS = [
    "Template",
    "Extraction",
    "MC_Note_Type",
    "File",
    "Operation Number",
    "Validation Date",
    "Author",
    "Page_Count",
    "Page count before opinion",
    "Page Count Annex",
    "Text Before Opinions",
]

DEFAULT_PREFIX_MERGE = {
    "OPS": "OPS",
    "OPS DIR": "OPS",
    "OPS/GLO": "OPS",
    "GLO": "GLO",
    "PJ": "PJ",
    "PJ DIR": "PJ",
    "RM": "RM",
    "GR&C-RM": "RM",
    "GR_C-RM": "RM",
    "GR&C RM": "RM",
    "GRC-RM": "RM",
    "GR C RM": "RM",
    "OCCO": "OCCO",
    "GR&C-OCCO": "OCCO",
    "GR_C-OCCO": "OCCO",
    "JU": "JU",
    "ECON": "ECON",
    "SG ECON": "ECON",
    "CFC": "CFC",
    "FC": "CFC",
    "EIF": "EIF",
    "FI": "FI",
    "IG": "IG",
    "IG/EV": "IG",
    "PMM": "PMM",
    "SG": "SG",
    "GIS": "GIS",
    "IS": "GIS",
    "HR": "HR",
    "OTHER": "OTHER",
}

START_SECTION_RE = re.compile(
    r"(6\s+OPINIONS\s+FROM\s+THE\s+SERVICES|OPINIONS\s+FROM\s+THE\s+SERVICES)",
    re.IGNORECASE,
)
OPINION_HEADER_RE = re.compile(r"\b\d+\.\d+\b.{0,80}\bOpinion\b", re.IGNORECASE)
END_BLOCK_RE = re.compile(
    r"\b[Vv]alidated\s+by\s+"
    r"[A-ZÀ-ÖØ-Þ][A-ZÀ-ÖØ-Þ'’\-.]{1,}"
    r"(?:\s+[A-ZÀ-ÖØ-Þ][A-Za-zÀ-ÖØ-öø-ÿ'’\-.]+)?"
)
STRICT_VALIDATION_BLOCK_RE = re.compile(
    r"\bValidated\s+by\s+.{2,220}?\bAt\s+"
    r"\d{2}(?:/|\.)\d{2}(?:/|\.)\d{4}\s+\d{1,2}:\d{2}\b",
    re.IGNORECASE | re.DOTALL,
)
LOOSE_VALIDATED_BY_SURNAME_RE = re.compile(r"\b[Vv]alidated\s+by\s+[A-ZÀ-ÖØ-Þ]{2,}\b")
WORD_RE = re.compile(r"\b[0-9A-Za-z']+\b")
MC_TYPE_RE = re.compile(r"(NOTEMCDEC|NOTEMCDISC|NOTEMCINFO)", re.IGNORECASE)
DATE_RE = re.compile(r"\b(20\d{2}-\d{2}-\d{2})\b")
AUTHOR_RE = re.compile(r"\b[A-Z]{1,5}(?:[/-][A-Z0-9&]{1,10})+\b")
DOC_DATE_RE = re.compile(r"\b(\d{2}/\d{2}/\d{4})\b")
ISO_DATE_RE = re.compile(r"\b(20\d{2}-\d{2}-\d{2})\b")
AUTHOR_LINE_RE = re.compile(r"^[A-ZÀ-ÖØ-Þ]{2,}[A-Za-z0-9À-ÖØ-öø-ÿ&/.,()'\- ]{3,}$")
OPERATION_NUMBER_LINE_RE = re.compile(r"^\s*(?:\d{8}|\d{4}-\d{4,6})\s*$")
EMBEDDED_OPERATION_NUMBER_RE = re.compile(r"(?<![A-Za-z0-9])20\d{2}-(?:\d{0,6})(?![A-Za-z0-9])")
OPERATION_NUMBER_RE = re.compile(r"(?<![A-Za-z0-9])(?:20\d{6}|20\d{2}-\d{4,6})(?![A-Za-z0-9])")


def normalize_for_match(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def clean_cell_value(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def repair_mojibake_text(value: str) -> str:
    text = str(value or "")
    replacements = {
        "\u00e2\u20ac\u201c": "\u2013",
        "\u00e2\u20ac\u0093": "\u2013",
        "\u00e2\u20ac\u201d": "\u2014",
        "\u00e2\u20ac\u0094": "\u2014",
        "\u00e2\u20ac\u02dc": "\u2018",
        "\u00e2\u20ac\u2122": "\u2019",
        "\u00e2\u20ac\u0153": "\u201c",
        "\u00e2\u20ac\u009d": "\u201d",
        "\u00e2\u20ac\u00a6": "\u2026",
        "\u00c2\u00a0": " ",
    }
    for bad, good in replacements.items():
        text = text.replace(bad, good)

    for _ in range(2):
        if not any(marker in text for marker in ("\u00c3", "\u00c2", "\u00e2", "\u20ac", "\ufffd")):
            break
        try:
            fixed = text.encode("cp1252").decode("utf-8")
        except UnicodeError:
            break
        if fixed == text:
            break
        text = fixed

    return unicodedata.normalize("NFC", text)


def clean_export_value(value):
    if isinstance(value, str):
        return repair_mojibake_text(value)
    return value


def format_iso_date_ddmmyyyy(value: str) -> str:
    match = re.fullmatch(r"(20\d{2})-(\d{2})-(\d{2})", str(value or "").strip())
    if not match:
        return ""
    yyyy, mm, dd = match.groups()
    return f"{dd}/{mm}/{yyyy}"


def normalize_validation_date(value: str) -> str:
    text = str(value or "").strip()
    if not text:
        return ""

    iso_date = format_iso_date_ddmmyyyy(text)
    if iso_date:
        return iso_date

    match = re.fullmatch(r"(\d{2})[/.](\d{2})[/.](20\d{2})", text)
    if match:
        dd, mm, yyyy = match.groups()
        return f"{dd}/{mm}/{yyyy}"

    return text


def header_key(value) -> str:
    return normalize_for_match(value)


def clean_text(text: str):
    text = str(text or "").replace("\u200b", "")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def clean_author_line(line: str) -> str:
    line = clean_cell_value(line)
    if not line:
        return ""
    line = EMBEDDED_OPERATION_NUMBER_RE.sub("", line)
    line = re.sub(r"/{2,}", "/", line)
    line = re.sub(r"\s+", " ", line).strip()
    return re.sub(r"(?<!-)/$", "", line).strip()


def count_words(text: str):
    return len(WORD_RE.findall(str(text or "")))


def open_pdf(pdf_path: Path):
    filename = str(Path(pdf_path).resolve())
    if len(filename) >= 240 and not filename.startswith("\\\\?\\"):
        filename = f"\\\\?\\{filename}"
    return fitz.open(filename)


def canonicalize_merge_to(value: str) -> str:
    value = clean_cell_value(value)
    output_lookup = {col.upper(): col for col in OUTPUT_PREFIXES}
    return output_lookup.get(value.upper(), value)


def build_normalized_prefixes(prefix_to_group):
    normalized_prefixes = []
    for prefix in prefix_to_group:
        norm = normalize_for_match(prefix)
        if norm:
            normalized_prefixes.append((norm, prefix))
    normalized_prefixes.sort(key=lambda item: len(item[0]), reverse=True)
    return normalized_prefixes


def find_prefix_merge_file(folder_path: Path) -> Path | None:
    allowed_suffixes = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    for suffix in allowed_suffixes:
        candidate = folder_path / f"{PREFIX_MERGE_BASENAME}{suffix}"
        if candidate.exists() and candidate.is_file():
            return candidate
    for file_path in folder_path.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in allowed_suffixes:
            if file_path.stem.upper() == PREFIX_MERGE_BASENAME.upper():
                return file_path
    return None


def load_prefix_merge_mapping(folder_path: Path):
    mapping_file = find_prefix_merge_file(folder_path)
    warnings = []

    if mapping_file is None:
        prefix_to_group = dict(DEFAULT_PREFIX_MERGE)
        warnings.append("PREFIX_MERGE workbook not found; used built-in default prefix mapping.")
        return None, prefix_to_group, build_normalized_prefixes(prefix_to_group), warnings

    if load_workbook is None:
        raise SystemExit("Missing dependency: openpyxl. Install it with: pip install openpyxl")

    workbook = load_workbook(mapping_file, data_only=True)
    worksheet = workbook.active
    prefix_col_idx = None
    merge_to_col_idx = None
    header_row_idx = None

    for row_idx in range(1, min(10, worksheet.max_row) + 1):
        row_values = [worksheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, worksheet.max_column + 1)]
        keys = [header_key(value) for value in row_values]
        if "PREFIX" in keys and "MERGETO" in keys:
            prefix_col_idx = keys.index("PREFIX") + 1
            merge_to_col_idx = keys.index("MERGETO") + 1
            header_row_idx = row_idx
            break

    if prefix_col_idx is None or merge_to_col_idx is None:
        raise ValueError(f"The file {mapping_file.name} must contain columns named PREFIX and MERGE TO.")

    prefix_to_group = {}
    for row_idx in range(header_row_idx + 1, worksheet.max_row + 1):
        prefix = clean_cell_value(worksheet.cell(row=row_idx, column=prefix_col_idx).value)
        merge_to = clean_cell_value(worksheet.cell(row=row_idx, column=merge_to_col_idx).value)
        if not prefix and not merge_to:
            continue
        if not prefix or not merge_to:
            warnings.append(f"Row {row_idx}: skipped incomplete mapping.")
            continue
        prefix_to_group[prefix] = canonicalize_merge_to(merge_to)

    if not prefix_to_group:
        raise ValueError(f"No valid prefix mappings were found in {mapping_file.name}.")

    return mapping_file, prefix_to_group, build_normalized_prefixes(prefix_to_group), warnings


def extract_text_from_pdf(pdf_path: Path):
    with open_pdf(pdf_path) as doc:
        page_texts = [page.get_text("text") or "" for page in doc]
    return "\n".join(page_texts), len(page_texts), page_texts


def extract_mc_metadata(filename: str):
    match_type = MC_TYPE_RE.search(filename)
    match_date = DATE_RE.search(filename)
    return (
        match_type.group(1).upper() if match_type else None,
        format_iso_date_ddmmyyyy(match_date.group(1)) if match_date else None,
    )


def extract_operation_number(filename: str, full_text: str = ""):
    for source in [filename, full_text[:4000]]:
        match = OPERATION_NUMBER_RE.search(source or "")
        if match:
            return match.group(0).replace("-", "")
    return ""


def is_valid_author_line(line: str) -> bool:
    line = clean_author_line(line)
    if not line:
        return False
    if line.lower().startswith("corporate use"):
        return False
    if OPERATION_NUMBER_LINE_RE.match(line):
        return False
    if "/" not in line and "-" not in line:
        return False
    return bool(AUTHOR_LINE_RE.match(line))


def extract_author_from_first_page(pdf_path: Path):
    with open_pdf(pdf_path) as doc:
        if len(doc) == 0:
            return None
        page1 = doc[0].get_text("text") or ""
    page1 = re.sub(r"\s+", " ", page1.replace("\u200b", " "))
    match = AUTHOR_RE.search(page1)
    if not match:
        return None
    candidate = clean_author_line(match.group(0))
    return candidate if is_valid_author_line(candidate) else None


def extract_top_right_text_page1(pdf_path: Path, x_frac=0.55, y_frac=0.40):
    with open_pdf(pdf_path) as doc:
        if len(doc) == 0:
            return ""
        page = doc[0]
        x_min = page.rect.width * x_frac
        y_max = page.rect.height * y_frac
        page_dict = page.get_text("dict")

    lines = []
    for block in page_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        bx0, by0, bx1, by1 = block.get("bbox", (0, 0, 0, 0))
        if bx1 < x_min or by0 > y_max:
            continue
        for line in block.get("lines", []):
            text = "".join(span.get("text", "") for span in line.get("spans", [])).strip()
            if text:
                y = line.get("bbox", (0, 0, 0, 0))[1]
                lines.append((y, clean_text(text)))

    lines.sort(key=lambda item: item[0])
    output = []
    previous = None
    for _, text in lines:
        if text != previous:
            output.append(text)
        previous = text
    return "\n".join(output).strip()


def extract_doc_date_and_author_top_right(pdf_path: Path):
    chunk = extract_top_right_text_page1(pdf_path)
    debug_info = {
        "top_right_chunk": chunk,
        "date_method": "",
        "date_trigger_text": "",
        "author_method": "",
        "author_trigger_text": "",
    }
    if not chunk:
        return None, None, debug_info

    lines = [line.strip() for line in chunk.splitlines() if line.strip()]
    doc_date = None
    date_idx = None
    for idx, line in enumerate(lines):
        match = DOC_DATE_RE.search(line)
        if match:
            doc_date = normalize_validation_date(match.group(1))
            date_idx = idx
            debug_info["date_method"] = "top-right DD/MM/YYYY"
            debug_info["date_trigger_text"] = line
            break

    if doc_date is None:
        for idx, line in enumerate(lines):
            match = ISO_DATE_RE.search(line)
            if match:
                doc_date = format_iso_date_ddmmyyyy(match.group(1))
                date_idx = idx
                debug_info["date_method"] = "top-right ISO date converted"
                debug_info["date_trigger_text"] = line
                break

    author = None
    search_lines = lines[date_idx + 1:] if date_idx is not None else lines
    for raw_line in search_lines:
        cleaned = clean_author_line(raw_line)
        if is_valid_author_line(cleaned):
            author = cleaned
            debug_info["author_method"] = "valid cleaned service-code line"
            debug_info["author_trigger_text"] = raw_line
            break

    return doc_date, author, debug_info


def join_page_texts(page_texts):
    return "\n".join(page_texts)


def get_page_start_positions(page_texts):
    starts = []
    pos = 0
    for index, page_text in enumerate(page_texts):
        starts.append(pos)
        pos += len(page_text)
        if index < len(page_texts) - 1:
            pos += 1
    return starts


def get_page_number_for_abs_position(page_texts, abs_position: int):
    starts = get_page_start_positions(page_texts)
    for index, start in enumerate(starts):
        next_start = starts[index + 1] if index + 1 < len(starts) else None
        if next_start is None or start <= abs_position < next_start:
            return index + 1
    return len(starts)


def detect_prefix_from_header(header_text: str, normalized_prefixes):
    norm_header = normalize_for_match(header_text)
    for norm_prefix, original_prefix in normalized_prefixes:
        if norm_prefix in norm_header:
            return original_prefix
    return None


def prefix_is_near_start(header_text: str, prefix: str, max_norm_index: int = 18) -> bool:
    norm_header = normalize_for_match(header_text)
    norm_prefix = normalize_for_match(prefix)
    idx = norm_header.find(norm_prefix)
    return idx != -1 and idx <= max_norm_index


def get_annex_count_after_last_opinion(page_texts, prefix_to_group: dict, normalized_prefixes: list):
    total_pages = len(page_texts)
    full_text = join_page_texts(page_texts)
    start_match = START_SECTION_RE.search(full_text)
    debug_info = {
        "method": "after last accepted opinion block",
        "section_found": "YES" if start_match else "NO",
        "last_opinion_found": "NO",
        "last_opinion_end_page": "",
        "last_opinion_header": "",
        "last_opinion_end_trigger": "",
        "annex_count": NA_VALUE,
    }
    if not start_match:
        return NA_VALUE, debug_info

    parse_text = full_text[start_match.end():]
    base_offset = start_match.end()
    pos = 0
    last_opinion_end_page = None
    while True:
        header_match = OPINION_HEADER_RE.search(parse_text, pos)
        if not header_match:
            break
        header_text = parse_text[header_match.start():header_match.end()]
        prefix = detect_prefix_from_header(header_text, normalized_prefixes)
        if prefix is None or not prefix_is_near_start(header_text, prefix):
            pos = header_match.end()
            continue
        end_match = END_BLOCK_RE.search(parse_text, header_match.end())
        if not end_match:
            break
        block_text = clean_text(parse_text[header_match.end():end_match.start()])
        if count_words(block_text) >= 2:
            last_opinion_end_page = get_page_number_for_abs_position(page_texts, base_offset + end_match.end())
            debug_info.update({
                "last_opinion_found": "YES",
                "last_opinion_end_page": last_opinion_end_page,
                "last_opinion_header": header_text.strip(),
                "last_opinion_end_trigger": parse_text[end_match.start():end_match.end()].strip(),
            })
        pos = end_match.end()

    if last_opinion_end_page is None:
        return NA_VALUE, debug_info
    annex_count = total_pages - last_opinion_end_page
    debug_info["annex_count"] = annex_count
    return annex_count, debug_info


def get_annex_count_no_opinions_top_down(page_texts):
    total_pages = len(page_texts)
    debug_info = {
        "method": "top-down first strict/fallback validation marker",
        "marker_found": "NO",
        "marker_page": "",
        "marker_type": "",
        "marker_text": "",
        "annex_count": NA_VALUE,
    }
    for page_index, page_text in enumerate(page_texts):
        strict_match = STRICT_VALIDATION_BLOCK_RE.search(page_text or "")
        loose_match = LOOSE_VALIDATED_BY_SURNAME_RE.search(page_text or "")
        marker = strict_match or loose_match
        if marker:
            marker_page = page_index + 1
            annex_count = total_pages - marker_page
            debug_info.update({
                "marker_found": "YES",
                "marker_page": marker_page,
                "marker_type": "strict Validated by + timestamp" if strict_match else "fallback Validated by SURNAME",
                "marker_text": marker.group(0),
                "annex_count": annex_count,
            })
            return annex_count, debug_info
    return NA_VALUE, debug_info


def get_normal_annex_count_from_last_validated_by(page_texts):
    total_pages = len(page_texts)
    for page_index in range(total_pages - 1, -1, -1):
        if list(END_BLOCK_RE.finditer(page_texts[page_index] or "")):
            return total_pages - (page_index + 1)
    return NA_VALUE


def get_text_before_opinions(page_texts, page_count_annex):
    chunks = []
    for page_index, page_text in enumerate(page_texts):
        match = START_SECTION_RE.search(page_text)
        if match:
            chunks.append(page_text[:match.start()])
            text_before = "\n".join(chunks)
            text_before_clean = clean_text(text_before)
            return {
                "section_found": True,
                "fallback_used": False,
                "fallback_reason": "",
                "page_count_before_opinion": page_index + 1,
                "text_before_opinions": count_words(text_before_clean),
                "section_trigger_text": match.group(0),
                "section_page": page_index + 1,
                "section_offset": match.start(),
                "text_before_count_text": text_before_clean,
            }
        chunks.append(page_text)

    total_pages = len(page_texts)
    fallback_page_count = total_pages - page_count_annex if isinstance(page_count_annex, int) else total_pages
    fallback_page_count = max(0, min(fallback_page_count, total_pages))
    fallback_text_clean = clean_text("\n".join(page_texts[:fallback_page_count]))
    return {
        "section_found": False,
        "fallback_used": True,
        "fallback_reason": "Opinions section not found; used validation marker to exclude annex pages" if isinstance(page_count_annex, int) else "Opinions section not found; used full document",
        "page_count_before_opinion": fallback_page_count,
        "text_before_opinions": count_words(fallback_text_clean),
        "section_trigger_text": "",
        "section_page": "",
        "section_offset": "",
        "text_before_count_text": fallback_text_clean,
    }


def parse_opinions(text: str, prefix_to_group: dict, normalized_prefixes: list):
    parser_debug = {
        "section_found": False,
        "section_trigger_text": "",
        "headers_found": 0,
        "headers_accepted": 0,
        "headers_rejected": 0,
        "header_candidates": [],
        "block_count": 0,
    }
    start_match = START_SECTION_RE.search(text)
    if not start_match:
        return None, None, parser_debug

    parser_debug["section_found"] = True
    parser_debug["section_trigger_text"] = start_match.group(0)
    text = text[start_match.end():]
    prefix_counts = defaultdict(int)
    debug_blocks = defaultdict(list)
    pos = 0

    while True:
        header_match = OPINION_HEADER_RE.search(text, pos)
        if not header_match:
            break
        parser_debug["headers_found"] += 1
        header_text = text[header_match.start():header_match.end()]
        prefix = detect_prefix_from_header(header_text, normalized_prefixes)
        if prefix is None or not prefix_is_near_start(header_text, prefix):
            parser_debug["headers_rejected"] += 1
            pos = header_match.end()
            continue
        end_match = END_BLOCK_RE.search(text, header_match.end())
        if not end_match:
            parser_debug["headers_rejected"] += 1
            break
        block_text = clean_text(text[header_match.end():end_match.start()])
        word_count = count_words(block_text)
        if word_count < 2:
            parser_debug["headers_rejected"] += 1
            pos = end_match.end()
            continue
        parser_debug["headers_accepted"] += 1
        parser_debug["block_count"] += 1
        prefix_counts[prefix] += word_count
        debug_blocks[prefix].append({
            "words": word_count,
            "text": block_text,
            "trigger_start": header_text.strip(),
            "trigger_end": text[end_match.start():end_match.end()].strip(),
            "merge_to": prefix_to_group.get(prefix, ""),
        })
        pos = end_match.end()

    return prefix_counts, debug_blocks, parser_debug


def build_group_totals(prefix_counts: dict, prefix_to_group: dict):
    group_totals = defaultdict(int)
    for prefix, count in prefix_counts.items():
        merge_to = prefix_to_group.get(prefix)
        if merge_to:
            group_totals[merge_to] += count
    return dict(group_totals)


def build_base_row(pdf_file, mc_type, operation_number, validation_date, author, page_count, page_count_annex, text_before_debug):
    return {
        "Template": TEMPLATE_VALUE,
        "Extraction": EXTRACTION_VALUE,
        "MC_Note_Type": mc_type if mc_type else NA_VALUE,
        "File": pdf_file.name,
        "Operation Number": operation_number,
        "Validation Date": normalize_validation_date(validation_date) if validation_date else NA_VALUE,
        "Author": author if author else NA_VALUE,
        "Page_Count": page_count,
        "Page count before opinion": text_before_debug["page_count_before_opinion"],
        "Page Count Annex": page_count_annex,
        "Text Before Opinions": text_before_debug["text_before_opinions"],
    }


def process_pdf(pdf_file: Path, prefix_to_group: dict, normalized_prefixes: list):
    mc_type, filename_date = extract_mc_metadata(pdf_file.name)
    raw_text, page_count, page_texts = extract_text_from_pdf(pdf_file)
    cleaned_text = clean_text(raw_text)
    operation_number = extract_operation_number(pdf_file.name, raw_text)

    no_opinion_annex_count, no_opinion_annex_debug = get_annex_count_no_opinions_top_down(page_texts)
    text_before_debug = get_text_before_opinions(page_texts, no_opinion_annex_count)
    with_opinion_annex_count, with_opinion_annex_debug = get_annex_count_after_last_opinion(page_texts, prefix_to_group, normalized_prefixes)
    normal_annex_count = get_normal_annex_count_from_last_validated_by(page_texts)

    if text_before_debug["section_found"]:
        if isinstance(with_opinion_annex_count, int):
            page_count_annex = with_opinion_annex_count
            annex_logic_used = "with opinions: after last accepted opinion block"
        else:
            page_count_annex = normal_annex_count
            annex_logic_used = "with opinions: fallback to last normal Validated by page"
    else:
        page_count_annex = no_opinion_annex_count
        annex_logic_used = "without opinions: top-down first required validation marker"

    doc_date, hdr_author, metadata_debug = extract_doc_date_and_author_top_right(pdf_file)
    validation_date = doc_date if doc_date else filename_date
    author = hdr_author if hdr_author else extract_author_from_first_page(pdf_file)

    prefix_counts, debug_data, parser_debug = parse_opinions(cleaned_text, prefix_to_group, normalized_prefixes)
    row_base = build_base_row(pdf_file, mc_type, operation_number, validation_date, author, page_count, page_count_annex, text_before_debug)

    group_totals = {} if prefix_counts is None else build_group_totals(prefix_counts, prefix_to_group)
    return {
        "base_row": row_base,
        "section_missing": prefix_counts is None,
        "prefix_counts": prefix_counts,
        "group_totals": group_totals,
        "debug_data": debug_data,
        "metadata_debug": metadata_debug,
        "text_before_debug": text_before_debug,
        "parser_debug": parser_debug,
        "prefix_to_group": prefix_to_group,
        "with_opinion_annex_count": with_opinion_annex_count,
        "with_opinion_annex_debug": with_opinion_annex_debug,
        "no_opinion_annex_count": no_opinion_annex_count,
        "no_opinion_annex_debug": no_opinion_annex_debug,
        "normal_annex_count": normal_annex_count,
        "annex_logic_used": annex_logic_used,
    }


def build_final_csv_rows(processed_results, output_prefixes):
    rows = []
    for result in processed_results:
        row = dict(result["base_row"])
        for prefix in output_prefixes:
            row[prefix] = NA_VALUE if result["section_missing"] else result["group_totals"].get(prefix, "")
        rows.append(row)
    return rows


def update_group_values(group_values, result):
    prefix_counts = result.get("prefix_counts")
    if prefix_counts is None:
        return
    for prefix, count in prefix_counts.items():
        merge_to = result["prefix_to_group"].get(prefix)
        if merge_to and count > 0:
            group_values[merge_to].append(count)


def confidence_interval_95(data):
    if len(data) < 2:
        return (None, None)
    avg = statistics.mean(data)
    se = statistics.stdev(data) / math.sqrt(len(data))
    margin = 1.96 * se
    return (round(avg - margin, 2), round(avg + margin, 2))


def build_statistics_rows(output_prefixes, group_values):
    rows = []
    for group in output_prefixes:
        values = group_values.get(group, [])
        if not values:
            rows.append({"Group": group, "N": 0, "Mean": NA_VALUE, "Median": NA_VALUE, "Std_Dev": NA_VALUE, "Min": NA_VALUE, "Max": NA_VALUE, "CI95_Lower": NA_VALUE, "CI95_Upper": NA_VALUE})
            continue
        lower, upper = confidence_interval_95(values)
        rows.append({
            "Group": group,
            "N": len(values),
            "Mean": round(statistics.mean(values), 2),
            "Median": round(statistics.median(values), 2),
            "Std_Dev": round(statistics.stdev(values), 2) if len(values) >= 2 else NA_VALUE,
            "Min": min(values),
            "Max": max(values),
            "CI95_Lower": lower if lower is not None else NA_VALUE,
            "CI95_Upper": upper if upper is not None else NA_VALUE,
        })
    return rows


def write_csv(rows, output_csv: Path, fieldnames):
    cleaned_rows = [
        {key: clean_export_value(value) for key, value in row.items()}
        for row in rows
    ]
    with output_csv.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(cleaned_rows)


def write_csv_with_fallback(rows, output_csv: Path, fieldnames):
    try:
        write_csv(rows, output_csv, fieldnames)
        return output_csv
    except PermissionError:
        fallback = output_csv.with_name(f"{output_csv.stem}_new{output_csv.suffix}")
        write_csv(rows, fallback, fieldnames)
        print(f"WARNING: Could not overwrite {output_csv.resolve()} because it is open or locked.")
        print(f"WARNING: Wrote fallback CSV instead: {fallback.resolve()}")
        return fallback


def parse_args():
    parser = argparse.ArgumentParser(description="Extract OTHER MC note opinion word counts from PDFs.")
    parser.add_argument("folder", nargs="?", type=Path, default=FOLDER_PATH)
    parser.add_argument("--output-dir", type=Path, default=OUTPUT_DIR)
    parser.add_argument("--prefix-merge-dir", type=Path, default=PREFIX_MERGE_DIR)
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--status", action="store_true")
    return parser.parse_args()


def main():
    global DEBUG, STATUS_OUTPUT
    args = parse_args()
    DEBUG = args.debug
    STATUS_OUTPUT = args.debug or args.status

    output_dir = args.output_dir if args.output_dir.is_absolute() else Path.cwd() / args.output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    mapping_file, prefix_to_group, normalized_prefixes, mapping_warnings = load_prefix_merge_mapping(args.prefix_merge_dir)

    for warning in mapping_warnings:
        print(f"WARNING: {warning}")

    pdf_files = sorted(args.folder.rglob("*.pdf"))
    if not pdf_files:
        raise SystemExit(f"No PDFs found under: {args.folder.resolve()}")

    processed_results = []
    group_values = defaultdict(list)
    extra_output_prefixes_found = []

    for pdf_file in pdf_files:
        result = process_pdf(pdf_file, prefix_to_group, normalized_prefixes)
        processed_results.append(result)
        update_group_values(group_values, result)
        for group in result.get("group_totals", {}):
            if group not in OUTPUT_PREFIXES and group not in extra_output_prefixes_found:
                extra_output_prefixes_found.append(group)

    final_output_prefixes = OUTPUT_PREFIXES + extra_output_prefixes_found
    fieldnames = BASE_COLUMNS + final_output_prefixes
    rows = build_final_csv_rows(processed_results, final_output_prefixes)
    stats_rows = build_statistics_rows(final_output_prefixes, group_values)

    output_file = write_csv_with_fallback(rows, output_dir / "MCNOTES_opinion_word_counts.csv", fieldnames)
    stats_output_file = write_csv_with_fallback(
        stats_rows,
        output_dir / "MCNOTES_group_statistics.csv",
        ["Group", "N", "Mean", "Median", "Std_Dev", "Min", "Max", "CI95_Lower", "CI95_Upper"],
    )

    print("\n================ RUN SUMMARY ================")
    print(f"PDF files processed              : {len(pdf_files)}")
    print(f"CSV rows exported                : {len(rows)}")
    print(f"Extra output columns found       : {len(extra_output_prefixes_found)}")
    print(f"PDF folder used                  : {args.folder.resolve()}")
    print(f"Output folder used               : {output_dir.resolve()}")
    print(f"Prefix merge file used           : {mapping_file.resolve() if mapping_file else 'built-in default mapping'}")
    print(f"Individual file CSV completed    : {output_file.resolve()}")
    print(f"Statistics CSV completed         : {stats_output_file.resolve()}")
    print("=============================================")


if __name__ == "__main__":
    main()
